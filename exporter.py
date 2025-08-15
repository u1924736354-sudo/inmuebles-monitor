import pandas as pd
from openpyxl.utils import get_column_letter

def export_to_excel(df, out_path):
    df.to_excel(out_path, index=False, sheet_name="Nuevos")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(out_path)
        ws = wb.active
        for i, col in enumerate(df.columns, start=1):
            maxlen = max(df[col].astype(str).map(len).max(), len(col)) + 2
            if maxlen > 60: maxlen = 60
            ws.column_dimensions[get_column_letter(i)].width = maxlen
        wb.save(out_path)
    except Exception as e:
        print("Excel formatting warning:", e)
    return out_path
